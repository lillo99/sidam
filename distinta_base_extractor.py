"""
Distinta Base Extractor - logica core (usabile da CLI e da Streamlit).
"""

import base64
import io
import json
import sys
from pathlib import Path

import anthropic
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

MEDIA_TYPES = {
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".gif": "image/gif",
    ".webp": "image/webp",
}

PROMPT = """Analizza questa schermata del gestionale che mostra una distinta base.
Estrai TUTTI i componenti visibili e restituisci un JSON con questa struttura esatta:

[
  {
    "denominazione": "descrizione completa del componente",
    "codice": "codice articolo (es. RC3082)",
    "quantita": "quantità in formato SCA/PZ (es. 1, 0,01, 1/10, 1/180, 1,15/180)",
    "um": "unità di misura (es. NR, MT)"
  }
]

IMPORTANTE:
- Per articoli con quantità intera usa "1" (non "1.0")
- Per quantità decimali usa la virgola come separatore (es. "0,01")
- Mantieni il formato frazionale dove presente (es. "1/180", "2/180")
- Restituisci SOLO il JSON, senza testo aggiuntivo."""


def extract_bom(image_bytes: bytes, media_type: str, api_key: str | None = None) -> list[dict]:
    """Invia l'immagine a Claude Vision e restituisce la lista degli articoli."""
    client = anthropic.Anthropic(api_key=api_key) if api_key else anthropic.Anthropic()
    image_data = base64.standard_b64encode(image_bytes).decode("utf-8")

    message = client.messages.create(
        model="claude-opus-4-6",
        max_tokens=2048,
        messages=[
            {
                "role": "user",
                "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": image_data}},
                    {"type": "text", "text": PROMPT},
                ],
            }
        ],
    )

    raw = message.content[0].text.strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    return json.loads(raw.strip())


def build_excel(items: list[dict], template_bytes: bytes | None = None) -> bytes:
    """
    Scrive gli articoli nel foglio DATABASE di un workbook.
    Se template_bytes è fornito, lo usa come base (preserva gli altri fogli).
    Restituisce i byte del file .xlsx.
    """
    if template_bytes:
        wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    if "DATABASE" in wb.sheetnames:
        ws = wb["DATABASE"]
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None
    else:
        ws = wb.create_sheet("DATABASE", 1)

    # Stili
    hdr_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    dat_font = Font(name="Arial", size=10)
    hdr_fill = PatternFill("solid", start_color="4472C4", end_color="4472C4")
    alt_fill = PatternFill("solid", start_color="DCE6F1", end_color="DCE6F1")
    thin = Side(border_style="thin", color="000000")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, header in enumerate(["Descrizione componente", "Codice articolo", "Quantità", "Unità di misura"], start=1):
        c = ws.cell(row=1, column=col, value=header)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = brd

    for r, item in enumerate(items, start=2):
        fill = alt_fill if r % 2 == 0 else None
        for col, key in enumerate(["denominazione", "codice", "quantita", "um"], start=1):
            c = ws.cell(row=r, column=col, value=item.get(key, ""))
            c.font = dat_font
            c.alignment = Alignment(vertical="center", wrap_text=(col == 1))
            c.border = brd
            if fill:
                c.fill = fill

    ws.column_dimensions["A"].width = 70
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 8
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Uso: python distinta_base_extractor.py <immagine> [<template.xlsx>]")
        sys.exit(1)

    image_path = Path(sys.argv[1])
    template_bytes = Path(sys.argv[2]).read_bytes() if len(sys.argv) > 2 else None
    media_type = MEDIA_TYPES.get(image_path.suffix.lower(), "image/png")

    print(f"Analisi: {image_path.name}")
    items = extract_bom(image_path.read_bytes(), media_type)
    print(f"Estratti {len(items)} componenti:")
    for item in items:
        print(f"  {item['codice']:10} | {item['quantita']:10} {item['um']:3} | {item['denominazione'][:60]}")

    out_bytes = build_excel(items, template_bytes)
    out_path = image_path.with_suffix(".xlsx")
    out_path.write_bytes(out_bytes)
    print(f"\nSalvato: {out_path}")


if __name__ == "__main__":
    main()
